import random
import time
from feasibility import is_feasible
from distance_finder import calculate_total_distance, distance_matrix_generator
from file_writer import save_to_excel
from file_reader import read_txt_file, Node
from solution_interpreter import info_of_all_routes
from openpyxl import Workbook
from gap_calculator import read_lower_bounds, write_GAP_excel
from vnd import vnd_algorithm

alpha = 1
beta = 1000


def genetic_algorithm(initial_routes, times, Q, remaining_time, start_time):
    """
    Implementación del algoritmo genético para mejorar la solución.
    """
    population_size = 1000
    crossover_rate = 0.9
    mutation_rate = 0.2
    generations = 20000

    # Generar la población inicial
    population = generate_initial_population(initial_routes, population_size)

    # Evaluar la población inicial
    fitness_values = evaluate_population(population, times, Q)

    # Ejecutar el ciclo de generaciones del algoritmo genético
    for generation in range(generations):
        # Verificar si se ha excedido el tiempo límite
        if time.time() - start_time > remaining_time:
            print(f"Tiempo límite alcanzado en la generación {generation}")
            break

        # Seleccionar padres para el cruce
        parents = select_parents(population, fitness_values)

        # Realizar el cruce
        offspring = crossover(parents, crossover_rate)

        # Aplicar mutación
        mutated_offspring = mutate(offspring, mutation_rate)

        # Filtrar la descendencia y quedarnos solo con las rutas factibles
        valid_offspring = []
        for child in mutated_offspring:
            if all(isinstance(node, Node) for node in child):  # Verificar que todos los elementos sean nodos
                if is_feasible(child, Q, times):
                    valid_offspring.append(child)

        # Si no hay descendencia válida, continuar con la siguiente generación
        if not valid_offspring:
            continue

        # Evaluar la descendencia válida
        offspring_fitness = evaluate_population(valid_offspring, times, Q)

        # Seleccionar sobrevivientes
        population = select_survivors(population, valid_offspring, offspring_fitness, times, Q)

        # Evaluar la nueva población
        fitness_values = evaluate_population(population, times, Q)

    # Obtener la mejor solución encontrada
    best_solution = get_best_solution(population, fitness_values)
    return best_solution


def generate_initial_population(initial_routes, population_size):
    """
    Genera una población inicial de rutas aleatorias basadas en la solución inicial.
    """
    population = []
    for _ in range(population_size):
        # Realiza una copia aleatoria de la solución inicial o variaciones de ella
        individual = random.sample(initial_routes, len(initial_routes))
        population.append(individual)
    return population

def evaluate_population(population, times, Q):
    """
    Evalúa la población calculando la aptitud (fitness) de cada individuo.
    Cuanto menor sea el costo, mejor es la solución.
    """
    fitness_values = []
    epsilon = 1e-6  # Pequeño valor para evitar división por cero y problemas numéricos
    
    # Calcular el total_cost para cada individuo
    total_costs = []
    for individual in population:
        total_cost = calculate_total_cost(individual, times, alpha=alpha, beta=beta)  # Ajusta alpha y beta según sea necesario
        total_costs.append(total_cost)
    
    # Encontrar el costo máximo y mínimo para normalizar
    max_cost = max(total_costs)
    min_cost = min(total_costs)
    cost_range = max_cost - min_cost + epsilon  # Añadimos epsilon para evitar división por cero si max_cost == min_cost
    
    # Calcular el fitness normalizado inverso (mejor costo => mayor fitness)
    for total_cost in total_costs:
        # Normalizar el costo entre 0 y 1
        normalized_cost = (total_cost - min_cost) / cost_range
        # Invertir el costo normalizado para que menor costo dé mayor fitness
        fitness = 1 / (normalized_cost + epsilon)
        fitness_values.append(fitness)

    return fitness_values


def calculate_total_cost(routes, times, alpha=1.0, beta=10000.0):
    total_distance = calculate_total_distance(routes, times)
    num_routes = len(routes)
    total_cost = alpha * total_distance + beta * num_routes
    return total_cost


def select_parents(population, fitness_values):
    """
    Selecciona a los padres utilizando el método de ruleta o selección por torneo.
    """
    total_fitness = sum(fitness_values)
    selection_probs = [fitness / total_fitness for fitness in fitness_values]
    parents = random.choices(population, weights=selection_probs, k=2)  # Selecciona 2 padres
    return parents

def crossover(parents, crossover_rate):
    """
    Realiza el cruce de dos padres para generar descendencia.
    """
    if random.random() < crossover_rate:
        # Cruza dos padres y genera un hijo
        crossover_point = random.randint(1, len(parents[0]) - 1)
        child1 = parents[0][:crossover_point] + [x for x in parents[1] if x not in parents[0][:crossover_point]]
        child2 = parents[1][:crossover_point] + [x for x in parents[0] if x not in parents[1][:crossover_point]]
        return [child1, child2]
    else:
        # No hay cruce, los padres pasan directamente a la siguiente generación
        return parents

def mutate(offspring, mutation_rate):
    """
    Realiza una mutación aleatoria en la descendencia.
    """
    for child in offspring:
        if random.random() < mutation_rate:
            # Realiza una mutación al cambiar el orden de dos rutas
            idx1, idx2 = random.sample(range(len(child)), 2)
            child[idx1], child[idx2] = child[idx2], child[idx1]
    return offspring

def get_best_solution(population, fitness_values):
    """
    Devuelve la mejor solución de la población, priorizando primero el menor fitness (distancia total) y, 
    en caso de empate, el menor número de rutas.
    """
    # Validaciones
    if not population or not fitness_values or len(population) != len(fitness_values):
        raise ValueError("La población y los valores de fitness deben ser listas no vacías de igual longitud.")
    
    # Encontrar el mínimo valor de fitness
    min_fitness = min(fitness_values)
    # Indices de individuos con el mínimo fitness
    best_indices = [i for i, fitness in enumerate(fitness_values) if fitness == min_fitness]
    
    if len(best_indices) == 1:
        # Solo hay un individuo con el mejor fitness
        best_idx = best_indices[0]
    else:
        # Si hay múltiples soluciones óptimas en términos de fitness, considerar el número de rutas
        min_routes = min(len(population[i]) for i in best_indices)
        # Indices de individuos con el menor número de rutas entre los mejores fitness
        candidates = [i for i in best_indices if len(population[i]) == min_routes]
        
        if len(candidates) == 1:
            # Solo hay un individuo con el mejor fitness y menor número de rutas
            best_idx = candidates[0]
        else:
            # Si aún hay empate, seleccionamos el primer candidato
            best_idx = candidates[0]
    
    return population[best_idx]

def select_survivors(population, offspring, fitness_values, times, Q):
    """
    Selecciona a los mejores individuos (padres + descendencia) para formar la nueva población.
    """
    combined_population = population + offspring
    combined_fitness = evaluate_population(combined_population, times, Q)
    
    # Selecciona los mejores individuos (top k) para sobrevivir
    survivors_idx = sorted(range(len(combined_fitness)), key=lambda i: combined_fitness[i], reverse=True)[:len(population)]
    survivors = [combined_population[i] for i in survivors_idx]
    return survivors


## Constructive method to select the "optimal" route based on the above restrictions
def humble_constructive(nodes, capacity, times):
    depot = nodes[0]
    customers = nodes[1:]
    routes = []
    while customers:
        route = [depot]
        current_load = 0
        while True:
            feasible_customers = [cust for cust in customers if is_feasible(route, capacity, times)]
            if not feasible_customers:
                break
            next_customer = min(feasible_customers, key=lambda x: times[route[-1].index][x.index])
            if current_load + next_customer.q <= capacity:
                route.append(next_customer)
                current_load += next_customer.q
                customers.remove(next_customer)
            else:
                break
        route.append(depot)
        routes.append(route)
    return routes


def get_initial_solution(method, nodes, Q, distances, initial_solution_path, sheet_name):
    # Obtener la solución inicial con adaptación de datos similar a `constructive`
    if method == "humble":
        return humble_constructive(nodes, Q, distances)
    initial_solution, _, _ = info_of_all_routes(initial_solution_path, sheet_name)
    for route_data in initial_solution:
        route_data['route_objects'] = [nodes[node_index] for node_index in route_data['route_indexes']]
    return [i['route_objects'] for i in initial_solution]

def get_initial_solution(method, nodes, Q, distances, initial_solution_path, sheet_name):
    # Obtener la solución inicial con adaptación de datos similar a `constructive`
    if method == "humble":
        return humble_constructive(nodes, Q, distances)
    initial_solution, _, _ = info_of_all_routes(initial_solution_path, sheet_name)
    for route_data in initial_solution:
        route_data['route_objects'] = [nodes[node_index] for node_index in route_data['route_indexes']]
    return [i['route_objects'] for i in initial_solution]

# Main execution
if __name__ == "__main__":
    initial_methods = ['constructive', 'GRASP', 'ACO', 'humble']
    vnd_options = [True, False]  # Opciones para aplicar VND
    current_method = 'GA'
    folder_name = '4-evolutionary-methods'
    instances_directory_path = 'VRPTW Instances'
    LB_file_directory = 'VRPTW Instances/LB_VRPTW.xlsx'

    LB_K, LB_D = read_lower_bounds(LB_file_directory, 'Hoja1')

    all_18_routes = {(method, vnd): [] for method in initial_methods for vnd in vnd_options}
    computation_times = {(method, vnd): [] for method in initial_methods for vnd in vnd_options}
    D = {(method, vnd): [] for method in initial_methods for vnd in vnd_options}
    K = {(method, vnd): [] for method in initial_methods for vnd in vnd_options}

    for initial_method in initial_methods:
        for apply_vnd in vnd_options:
            print(f"\nProcessing for initial method: {initial_method} with VND: {apply_vnd}")

            # Crear un archivo Excel para cada combinación de método y VND
            results_excel_path = f"{folder_name}/results/VRPTW_tm_{current_method}_ini_{initial_method}_VND_{apply_vnd}.xlsx"
            wb_results = Workbook()
            if "Sheet" in wb_results.sheetnames:
                wb_results.remove(wb_results["Sheet"])  # Eliminar hoja inicial vacía

            for sheet_number in range(1, 19):
                sheet_name = f'VRPTW{sheet_number}'
                instance_filename = f'{instances_directory_path}/{sheet_name}.txt'

                # Configuración de tiempo restante
                if sheet_number >= 1 and sheet_number <= 6:
                    remaining_time = 50e3
                elif sheet_number >= 7 and sheet_number <= 12:
                    remaining_time = 200e3
                else:
                    remaining_time = 750e3

                # Leer datos de entrada
                n, Q, nodes = read_txt_file(instance_filename)
                distances = distance_matrix_generator(nodes)

                print(f" - Processing {sheet_name} with initial method: {initial_method}, VND: {apply_vnd}")
                start_time = time.time()
                initial_solution_path = f'{folder_name}/constructive-results/VRPTW_tm_{initial_method}.xlsx' if initial_method != 'humble' else None

                # Obtener la solución inicial
                initial_solution = get_initial_solution(initial_method, nodes, Q, distances, initial_solution_path, sheet_name)

                # Ejecutar el Algoritmo Genético
                best_solution = genetic_algorithm(initial_solution, distances, Q, remaining_time, start_time)

                # Verificar si aún queda tiempo para ejecutar VND
                elapsed_time = time.time() - start_time
                if apply_vnd and elapsed_time < remaining_time:
                    # Ejecutar VND si se habilitó y queda tiempo
                    vnd_best_solution = vnd_algorithm(best_solution, distances, Q, remaining_time - elapsed_time, start_time=time.time())
                    final_solution = vnd_best_solution
                    computation_time = time.time() - start_time
                else:
                    # Usar solo la solución del GA sin VND
                    final_solution = best_solution
                    computation_time = elapsed_time

                # Calcular el número de rutas y la distancia total
                total_distance = calculate_total_distance(final_solution, distances)
                route_count = len(final_solution)

                print(f"   - Total Distance: {total_distance}")
                print(f"   - Number of Routes: {route_count}")
                print(f"   - Computation Time: {computation_time:.2f} seconds")

                # Guardar resultados en diccionarios
                all_18_routes[(initial_method, apply_vnd)].append(final_solution)
                computation_times[(initial_method, apply_vnd)].append(computation_time)
                D[(initial_method, apply_vnd)].append(total_distance)
                K[(initial_method, apply_vnd)].append(route_count)

                # Guardar la hoja en el archivo Excel de resultados
                save_to_excel(wb_results, sheet_name, final_solution, total_distance, computation_time, distances)

            # Guardar el archivo Excel al final de todas las instancias
            wb_results.save(results_excel_path)

    # Crear un único archivo para los GAPs
    wb_gaps = Workbook()
    if "Sheet" in wb_gaps.sheetnames:
        wb_gaps.remove(wb_gaps["Sheet"])

    for method in initial_methods:
        for vnd in vnd_options:
            sheet_name = f"{method}_VND_{vnd}"
            write_GAP_excel(wb_gaps, LB_K, K[(method, vnd)], LB_D, D[(method, vnd)], computation_times[(method, vnd)], sheet_name)

    wb_gaps.save(f"{folder_name}/results/GAPs_for_{current_method}_with_all_methods.xlsx")

