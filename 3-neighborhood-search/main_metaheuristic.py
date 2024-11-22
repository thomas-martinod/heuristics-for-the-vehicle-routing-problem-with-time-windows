import os
import math
import numpy as np
import time
from feasibility import is_feasible
from openpyxl import load_workbook, Workbook
from file_writer import save_to_excel
from file_reader import Node, read_txt_file
from distance_finder import calculate_total_distance, distance_matrix_generator
from gap_calculator import read_lower_bounds, write_GAP_excel
from solution_interpreter import info_of_all_routes
from vnd import vnd_algorithm
import random

# Global parameters
alpha = 1
beta = 1000
MAX_NO_IMPROVEMENT = 500  # Número máximo de iteraciones sin mejora


# Simulated annneaing parameters
initial_temperature = 300
cooling_rate = 0.95


# Tabu search parameters
tabu_tenure = 10



def get_initial_solution(method, nodes, Q, distances, initial_solution_path, sheet_name):
    initial_solution, _, _ = info_of_all_routes(initial_solution_path, sheet_name)
    for route_data in initial_solution:
        route_data['route_objects'] = [nodes[node_index] for node_index in route_data['route_indexes']]
    return [i['route_objects'] for i in initial_solution]


def calculate_total_cost(routes, times, alpha=alpha, beta=beta):
    total_distance = calculate_total_distance(routes, times)
    num_routes = len(routes)
    total_cost = alpha * total_distance + beta * num_routes
    return total_cost


def generate_neighbor(current_routes, times, capacity):
    neighbor_routes = [route.copy() for route in current_routes]
    move_type = random.choice(['swap_within_route', 'merge_routes', 'relocate_customer'])

    if move_type == 'swap_within_route':
        route_idx = random.randint(0, len(neighbor_routes) - 1)
        route = neighbor_routes[route_idx]
        if len(route) > 3:
            i = random.randint(1, len(route) - 3)
            j = random.randint(i + 1, len(route) - 2)
            new_route = route[:i] + route[i:j+1][::-1] + route[j+1:]
            if is_feasible(new_route, capacity, times):
                neighbor_routes[route_idx] = new_route

    elif move_type == 'merge_routes':
        if len(neighbor_routes) > 1:
            idx1, idx2 = random.sample(range(len(neighbor_routes)), 2)
            route1 = neighbor_routes[idx1]
            route2 = neighbor_routes[idx2]
            merged_route = route1[:-1] + route2[1:]
            if is_feasible(merged_route, capacity, times):
                neighbor_routes.pop(max(idx1, idx2))
                neighbor_routes.pop(min(idx1, idx2))
                neighbor_routes.append(merged_route)

    elif move_type == 'relocate_customer':
        from_route_idx = random.randint(0, len(neighbor_routes) - 1)
        to_route_idx = random.randint(0, len(neighbor_routes) - 1)
        while from_route_idx == to_route_idx:
            to_route_idx = random.randint(0, len(neighbor_routes) - 1)
        from_route = neighbor_routes[from_route_idx]
        to_route = neighbor_routes[to_route_idx]
        if len(from_route) > 3:
            customer_idx = random.randint(1, len(from_route) - 2)
            customer = from_route.pop(customer_idx)
            insertion_idx = random.randint(1, len(to_route) - 1)
            new_to_route = to_route[:insertion_idx] + [customer] + to_route[insertion_idx:]
            if is_feasible(new_to_route, capacity, times) and is_feasible(from_route, capacity, times):
                neighbor_routes[from_route_idx] = from_route
                neighbor_routes[to_route_idx] = new_to_route
            else:
                from_route.insert(customer_idx, customer)

    return neighbor_routes

def generate_neighborhood(current_routes, times, capacity, neighborhood_size=10):
    neighborhood = []
    for _ in range(neighborhood_size):
        neighbor = generate_neighbor(current_routes, times, capacity)
        neighborhood.append(neighbor)
    return neighborhood

def extract_move(current_routes, candidate_routes):
    current_routes_set = set(frozenset(route) for route in current_routes)
    candidate_routes_set = set(frozenset(route) for route in candidate_routes)
    move = current_routes_set.symmetric_difference(candidate_routes_set)
    return move

def tabu_search_dynamic(routes, times, capacity, initial_tabu_tenure, time_limit, start_time, alpha=1.0, beta=500.0):
    best_routes = [route.copy() for route in routes]
    best_cost = calculate_total_cost(best_routes, times, alpha, beta)
    current_routes = best_routes.copy()
    current_cost = best_cost
    tabu_list = []
    tabu_tenure = initial_tabu_tenure
    max_tabu_size = initial_tabu_tenure
    no_improvement_counter = 0

    initial_neighborhood_size = 100  # Tamaño inicial del vecindario
    neighborhood_size = initial_neighborhood_size  # Ajustaremos dinámicamente este tamaño

    while no_improvement_counter < MAX_NO_IMPROVEMENT * 2:
        current_time = time.time()
        if current_time - start_time >= time_limit:
            break

        # Generamos el vecindario con el tamaño dinámico
        neighborhood = generate_neighborhood(current_routes, times, capacity, neighborhood_size)
        best_candidate = None
        best_candidate_cost = float('inf')
        best_move = None

        # Buscamos el mejor candidato en el vecindario
        for candidate in neighborhood:
            candidate_cost = calculate_total_cost(candidate, times, alpha, beta)
            move = extract_move(current_routes, candidate)

            # Permitir movimientos tabu si hay estancamiento o si mejora significativamente
            if move not in tabu_list or candidate_cost < best_cost * 1.05:  # Permitir movimiento tabu si mejora significativamente
                if candidate_cost < best_candidate_cost:
                    best_candidate = candidate
                    best_candidate_cost = candidate_cost
                    best_move = move

        if best_candidate is None:
            break

        # Actualizamos la solución actual y la lista tabu
        current_routes = best_candidate
        current_cost = best_candidate_cost

        # Si encontramos una mejor solución, actualizamos el mejor registro
        if best_candidate_cost < best_cost:
            best_routes = best_candidate
            best_cost = best_candidate_cost
            no_improvement_counter = 0
            neighborhood_size = initial_neighborhood_size  # Restauramos el tamaño del vecindario
        else:
            no_improvement_counter += 1

        # Actualizamos la lista tabu con manejo dinámico de tabu_tenure
        if best_move:
            tabu_list.append(best_move)
            if len(tabu_list) > max_tabu_size:
                tabu_list.pop(0)

        # Ajustes de parámetros dinámicos en función del progreso
        if no_improvement_counter > MAX_NO_IMPROVEMENT / 2:
            # Si estamos en estancamiento, aumentamos el tamaño del vecindario y el tabu_tenure para explorar
            neighborhood_size = int(initial_neighborhood_size * 1.5)
            tabu_tenure = min(tabu_tenure + 1, initial_tabu_tenure * 2)
            max_tabu_size = tabu_tenure
        else:
            # Si hay mejoras constantes, reducimos el vecindario y el tabu_tenure para intensificar
            neighborhood_size = max(initial_neighborhood_size, neighborhood_size - 1)
            tabu_tenure = max(1, tabu_tenure - 1)
            max_tabu_size = tabu_tenure

    return best_routes

def select_operator(operators, operator_scores):
    total_score = sum(operator_scores[op] for op in operators)
    pick = random.uniform(0, total_score)
    current = 0
    for op in operators:
        current += operator_scores[op]
        if current >= pick:
            return op
    return operators[-1]

def alns_algorithm(routes, times, capacity, destroy_operators, repair_operators, time_limit, start_time, alpha=10.0, beta=450.0):
    best_routes = [route.copy() for route in routes]
    best_cost = calculate_total_cost(best_routes, times, alpha, beta)
    current_routes = best_routes.copy()

    operator_scores = {op: 1 for op in destroy_operators + repair_operators}

    no_improvement_counter = 0

    while no_improvement_counter < MAX_NO_IMPROVEMENT:
        current_time = time.time()
        if current_time - start_time >= time_limit:
            break

        destroy_op = select_operator(destroy_operators, operator_scores)
        repair_op = select_operator(repair_operators, operator_scores)

        partial_routes, customers_to_reinsert = destroy_op(current_routes, times, capacity)
        new_routes = repair_op(partial_routes, customers_to_reinsert, times, capacity)
        new_cost = calculate_total_cost(new_routes, times, alpha, beta)

        if new_cost < best_cost:
            best_routes = new_routes
            best_cost = new_cost
            operator_scores[destroy_op] += 1
            operator_scores[repair_op] += 1
            no_improvement_counter = 0  # Reset counter
        else:
            operator_scores[destroy_op] = max(1, operator_scores[destroy_op] - 1)
            operator_scores[repair_op] = max(1, operator_scores[repair_op] - 1)
            no_improvement_counter += 1

        current_routes = new_routes

    return best_routes

def destroy_route_removal(routes, times, capacity):
    destroyed_routes = [route.copy() for route in routes]
    num_routes_to_remove = max(1, int(0.1 * len(routes)))
    routes_to_remove = random.sample(destroyed_routes, num_routes_to_remove)
    customers_to_reinsert = []

    for route in routes_to_remove:
        customers_to_reinsert.extend(route[1:-1])
        destroyed_routes.remove(route)
    return destroyed_routes, customers_to_reinsert

def destroy_least_utilized(routes, times, capacity):
    destroyed_routes = [route.copy() for route in routes]
    route_loads = [(sum(node.q for node in route[1:-1]), idx) for idx, route in enumerate(destroyed_routes)]
    route_loads.sort(key=lambda x: x[0])
    num_routes_to_remove = max(1, int(0.1 * len(routes)))
    routes_to_remove = [destroyed_routes[idx] for _, idx in route_loads[:num_routes_to_remove]]
    customers_to_reinsert = []
    for route in routes_to_remove:
        customers_to_reinsert.extend(route[1:-1])
        destroyed_routes.remove(route)
    return destroyed_routes, customers_to_reinsert

def repair_greedy(partial_routes, customers_to_insert, times, capacity):
    routes = [route.copy() for route in partial_routes]
    for customer in customers_to_insert:
        best_position = None
        best_increase = float('inf')
        best_route_idx = None
        for idx, route in enumerate(routes):
            for pos in range(1, len(route)):
                new_route = route[:pos] + [customer] + route[pos:]
                if is_feasible(new_route, capacity, times):
                    increase = times[route[pos - 1].index][customer.index] + times[customer.index][route[pos].index] - times[route[pos - 1].index][route[pos].index]
                    if increase < best_increase:
                        best_increase = increase
                        best_position = pos
                        best_route_idx = idx
        if best_position is not None:
            routes[best_route_idx] = routes[best_route_idx][:best_position] + [customer] + routes[best_route_idx][best_position:]
        else:
            new_route = [routes[0][0], customer, routes[0][0]]
            if is_feasible(new_route, capacity, times):
                routes.append(new_route)
    return routes

def repair_regret(partial_routes, customers_to_insert, times, capacity):
    routes = [route.copy() for route in partial_routes]
    while customers_to_insert:
        regrets = []
        for customer in customers_to_insert:
            insertion_costs = []
            for idx, route in enumerate(routes):
                best_increase = float('inf')
                for pos in range(1, len(route)):
                    new_route = route[:pos] + [customer] + route[pos:]
                    if is_feasible(new_route, capacity, times):
                        increase = times[route[pos - 1].index][customer.index] + times[customer.index][route[pos].index] - times[route[pos - 1].index][route[pos].index]
                        if increase < best_increase:
                            best_increase = increase
                if best_increase < float('inf'):
                    insertion_costs.append(best_increase)
            if len(insertion_costs) >= 2:
                insertion_costs.sort()
                regret = insertion_costs[1] - insertion_costs[0]
            elif len(insertion_costs) == 1:
                regret = insertion_costs[0]
            else:
                regret = float('inf')
            regrets.append((regret, customer))
        regrets.sort(key=lambda x: x[0], reverse=True)
        _, selected_customer = regrets[0]
        best_position = None
        best_increase = float('inf')
        best_route_idx = None
        for idx, route in enumerate(routes):
            for pos in range(1, len(route)):
                new_route = route[:pos] + [selected_customer] + route[pos:]
                if is_feasible(new_route, capacity, times):
                    increase = times[route[pos - 1].index][selected_customer.index] + times[selected_customer.index][route[pos].index] - times[route[pos - 1].index][route[pos].index]
                    if increase < best_increase:
                        best_increase = increase
                        best_position = pos
                        best_route_idx = idx
        if best_position is not None:
            routes[best_route_idx] = routes[best_route_idx][:best_position] + [selected_customer] + routes[best_route_idx][best_position:]
        else:
            new_route = [routes[0][0], selected_customer, routes[0][0]]
            if is_feasible(new_route, capacity, times):
                routes.append(new_route)
        customers_to_insert.remove(selected_customer)
    return routes

def repair_savings(partial_routes, customers_to_insert, times, capacity):
    routes = [route.copy() for route in partial_routes]
    depot = routes[0][0]
    savings = []
    for i in customers_to_insert:
        for j in customers_to_insert:
            if i != j:
                s = times[depot.index][i.index] + times[depot.index][j.index] - times[i.index][j.index]
                savings.append((s, i, j))
    savings.sort(key=lambda x: x[0], reverse=True)
    inserted_customers = set()
    for s, i, j in savings:
        if i in inserted_customers or j in inserted_customers:
            continue
        new_route = [depot, i, j, depot]
        if is_feasible(new_route, capacity, times):
            routes.append(new_route)
            inserted_customers.update([i, j])
    remaining_customers = [c for c in customers_to_insert if c not in inserted_customers]
    for customer in remaining_customers:
        best_position = None
        best_increase = float('inf')
        best_route_idx = None
        for idx, route in enumerate(routes):
            for pos in range(1, len(route)):
                new_route = route[:pos] + [customer] + route[pos:]
                if is_feasible(new_route, capacity, times):
                    increase = times[route[pos - 1].index][customer.index] + times[customer.index][route[pos].index] - times[route[pos - 1].index][route[pos].index]
                    if increase < best_increase:
                        best_increase = increase
                        best_position = pos
                        best_route_idx = idx
        if best_position is not None:
            routes[best_route_idx] = routes[best_route_idx][:best_position] + [customer] + routes[best_route_idx][best_position:]
        else:
            new_route = [depot, customer, depot]
            if is_feasible(new_route, capacity, times):
                routes.append(new_route)
    return routes


def destroy_random(routes, times, capacity):
    destroyed_routes = [route.copy() for route in routes]
    num_customers_to_remove = max(1, int(0.1 * sum(len(route) - 2 for route in routes)))
    customers_to_remove = []
    while len(customers_to_remove) < num_customers_to_remove:
        route_idx = random.randint(0, len(destroyed_routes) - 1)
        route = destroyed_routes[route_idx]
        if len(route) > 3:
            cust_idx = random.randint(1, len(route) - 2)
            customer = route.pop(cust_idx)
            customers_to_remove.append(customer)
    return destroyed_routes, customers_to_remove

def destroy_worst(routes, times, capacity):
    destroyed_routes = [route.copy() for route in routes]
    customer_costs = []
    for route in destroyed_routes:
        for idx in range(1, len(route) - 1):
            prev_node = route[idx - 1]
            current_node = route[idx]
            next_node = route[idx + 1]
            cost = times[prev_node.index][current_node.index] + times[current_node.index][next_node.index] - times[prev_node.index][next_node.index]
            customer_costs.append((cost, current_node, route))
    customer_costs.sort(key=lambda x: x[0], reverse=True)
    num_customers_to_remove = max(1, int(0.1 * len(customer_costs)))
    customers_to_remove = []
    for i in range(num_customers_to_remove):
        cost, customer, route = customer_costs[i]
        route.remove(customer)
        customers_to_remove.append(customer)
    return destroyed_routes, customers_to_remove


def generate_multiple_neighbors(current_routes, times, capacity, num_neighbors=20):
    # Generamos varios vecinos y seleccionamos los más prometedores en términos de costo
    candidate_neighbors = []
    for _ in range(num_neighbors):  # Generamos un número elevado de vecinos para selección
        neighbor_routes = [route.copy() for route in current_routes]
        move_type = random.choice(['swap_within_route', 'merge_routes', 'relocate_customer'])

        if move_type == 'swap_within_route':
            route_idx = random.randint(0, len(neighbor_routes) - 1)
            route = neighbor_routes[route_idx]
            if len(route) > 3:
                i = random.randint(1, len(route) - 3)
                j = random.randint(i + 1, len(route) - 2)
                new_route = route[:i] + route[i:j+1][::-1] + route[j+1:]
                if is_feasible(new_route, capacity, times):
                    neighbor_routes[route_idx] = new_route

        elif move_type == 'merge_routes':
            if len(neighbor_routes) > 1:
                idx1, idx2 = random.sample(range(len(neighbor_routes)), 2)
                route1 = neighbor_routes[idx1]
                route2 = neighbor_routes[idx2]
                merged_route = route1[:-1] + route2[1:]
                if is_feasible(merged_route, capacity, times):
                    neighbor_routes.pop(max(idx1, idx2))
                    neighbor_routes.pop(min(idx1, idx2))
                    neighbor_routes.append(merged_route)

        elif move_type == 'relocate_customer':
            from_route_idx = random.randint(0, len(neighbor_routes) - 1)
            to_route_idx = random.randint(0, len(neighbor_routes) - 1)
            while from_route_idx == to_route_idx:
                to_route_idx = random.randint(0, len(neighbor_routes) - 1)
            from_route = neighbor_routes[from_route_idx]
            to_route = neighbor_routes[to_route_idx]
            if len(from_route) > 3:
                customer_idx = random.randint(1, len(from_route) - 2)
                customer = from_route.pop(customer_idx)
                insertion_idx = random.randint(1, len(to_route) - 1)
                new_to_route = to_route[:insertion_idx] + [customer] + to_route[insertion_idx:]
                if is_feasible(new_to_route, capacity, times) and is_feasible(from_route, capacity, times):
                    neighbor_routes[from_route_idx] = from_route
                    neighbor_routes[to_route_idx] = new_to_route
                else:
                    from_route.insert(customer_idx, customer)

        candidate_neighbors.append(neighbor_routes)

    # Seleccionamos los vecinos con menor costo
    candidate_neighbors.sort(key=lambda neighbor: calculate_total_cost(neighbor, times))
    return candidate_neighbors[:5]  # Retornamos los mejores 5 vecinos

def simulated_annealing_robust(routes, times, capacity, initial_temperature, cooling_rate, time_limit, start_time, alpha=alpha, beta=beta):
    best_routes = [route.copy() for route in routes]
    best_cost = calculate_total_cost(best_routes, times, alpha, beta)
    current_routes = best_routes.copy()
    current_cost = best_cost
    temperature = initial_temperature

    no_improvement_counter = 0
    max_no_improvement = MAX_NO_IMPROVEMENT * 2  # Aumentamos el límite de intentos sin mejora
    perturbation_chance = 0.2

    while temperature > 0.01 and no_improvement_counter < max_no_improvement:
        current_time = time.time()
        if current_time - start_time >= time_limit:
            break

        # Generamos múltiples vecinos y seleccionamos los mejores
        candidate_neighbors = generate_multiple_neighbors(current_routes, times, capacity)

        # Elegimos el mejor vecino
        for neighbor_routes in candidate_neighbors:
            neighbor_cost = calculate_total_cost(neighbor_routes, times, alpha, beta)
            delta = neighbor_cost - current_cost

            if delta < 0 or random.uniform(0, 1) < math.exp(-delta / temperature):
                current_routes = neighbor_routes
                current_cost = neighbor_cost

                # Si mejoramos el mejor costo global, lo actualizamos
                if neighbor_cost < best_cost:
                    best_routes = neighbor_routes
                    best_cost = neighbor_cost
                    no_improvement_counter = 0
                else:
                    no_improvement_counter += 1
            else:
                no_improvement_counter += 1

        # Forzamos una perturbación si estamos estancados
        if no_improvement_counter > MAX_NO_IMPROVEMENT and random.uniform(0, 1) < perturbation_chance:
            perturb_route = random.choice(current_routes)
            for _ in range(len(perturb_route) // 2):
                from_route_idx = random.randint(0, len(current_routes) - 1)
                to_route_idx = random.randint(0, len(current_routes) - 1)
                while from_route_idx == to_route_idx:
                    to_route_idx = random.randint(0, len(current_routes) - 1)
                from_route = current_routes[from_route_idx]
                to_route = current_routes[to_route_idx]
                if len(from_route) > 3:
                    customer_idx = random.randint(1, len(from_route) - 2)
                    customer = from_route.pop(customer_idx)
                    insertion_idx = random.randint(1, len(to_route) - 1)
                    to_route.insert(insertion_idx, customer)
            no_improvement_counter = 0  # Reset after perturbation

        # Ajustamos la temperatura y aplicamos reinicio si necesario
        temperature *= cooling_rate
        if no_improvement_counter > MAX_NO_IMPROVEMENT * 1.5:
            temperature = initial_temperature  # Reinicio de la temperatura

    return best_routes




if __name__ == "__main__":
    initial_methods = ['constructive', 'GRASP', 'ACO']
    folder_name = '3-neighborhood-search'
    instances_directory_path = 'VRPTW Instances'
    LB_file_directory = 'VRPTW Instances/LB_VRPTW.xlsx'

    computation_times = {method: [] for method in initial_methods}
    all_18_routes = {method: [] for method in initial_methods}
    D = {method: [] for method in initial_methods}
    K = {method: [] for method in initial_methods}

    destroy_operators = [destroy_random, destroy_worst, destroy_route_removal, destroy_least_utilized]
    repair_operators = [repair_greedy, repair_regret, repair_savings]

    # Leer los valores de referencia (Lower Bounds)
    LB_K, LB_D = read_lower_bounds(LB_file_directory, 'Hoja1')

    for initial_method in initial_methods:
        print(f"\nProcessing for initial method: {initial_method}")

        # Crear un archivo Excel para resultados
        results_excel_path = f"{folder_name}/results/VRPTW_tm_metaheuristic_ini_{initial_method}.xlsx"
        wb_results = Workbook()
        if "Sheet" in wb_results.sheetnames:
            wb_results.remove(wb_results["Sheet"])

        for sheet_number in range(1, 19):  # Iterar por todas las instancias
            sheet_name = f'VRPTW{sheet_number}'
            instance_filename = f'{instances_directory_path}/{sheet_name}.txt'

            # Configurar tiempo restante según la instancia
            if sheet_number >= 1 and sheet_number <= 6:
                remaining_time = 50e3
            elif sheet_number >= 7 and sheet_number <= 12:
                remaining_time = 200e3
            else:
                remaining_time = 750e3

            # Leer datos de entrada
            n, Q, nodes = read_txt_file(instance_filename)
            distances = distance_matrix_generator(nodes)

            print(f" - Processing {sheet_name} with initial method: {initial_method}")
            start_time = time.time()
            initial_solution_path = f'{folder_name}/constructive-results/VRPTW_tm_{initial_method}.xlsx'

            # Obtener solución inicial
            initial_solution = get_initial_solution(initial_method, nodes, Q, distances, initial_solution_path, sheet_name)

            # Ejecutar Metaheurísticas y optimización en paralelo
            routes = simulated_annealing_robust(
                initial_solution, distances, Q, initial_temperature, cooling_rate, 
                remaining_time, start_time, alpha, beta
            )
            elapsed_time = time.time() - start_time
            remaining_time -= elapsed_time

            if remaining_time > 0:
                routes = tabu_search_dynamic(
                    routes, distances, Q, tabu_tenure, remaining_time, time.time(), alpha, beta
                )
                elapsed_time = time.time() - start_time
                remaining_time -= elapsed_time

            if remaining_time > 0:
                routes = alns_algorithm(
                    routes, distances, Q, destroy_operators, repair_operators, 
                    remaining_time, time.time(), alpha, beta
                )
                elapsed_time = time.time() - start_time
                remaining_time -= elapsed_time

            # Aplicar VND si queda tiempo
            if remaining_time > 0:
                routes = vnd_algorithm(routes, distances, Q, remaining_time, time.time())

            # Calcular distancia total y número de rutas
            total_distance = calculate_total_distance(routes, distances)
            route_count = len(routes)

            print(f"   - Total Distance: {total_distance}")
            print(f"   - Number of Routes: {route_count}")
            print(f"   - Remaining Time: {remaining_time:.2f} ms")

            # Guardar resultados
            computation_times[initial_method].append(elapsed_time)
            D[initial_method].append(total_distance)
            K[initial_method].append(route_count)
            all_18_routes[initial_method].append(routes)

            # Guardar hoja en el Excel de resultados
            save_to_excel(wb_results, sheet_name, routes, total_distance, elapsed_time, distances)

        # Guardar archivo Excel al final
        wb_results.save(results_excel_path)

    # Crear archivo único para los GAPs
    wb_gaps = Workbook()
    if "Sheet" in wb_gaps.sheetnames:
        wb_gaps.remove(wb_gaps["Sheet"])

    for method in initial_methods:
        sheet_name = f"{method}"
        write_GAP_excel(wb_gaps, LB_K, K[method], LB_D, D[method], computation_times[method], sheet_name)

    wb_gaps.save(f"{folder_name}/results/GAPs_for_metaheuristic_with_all_methods.xlsx")
