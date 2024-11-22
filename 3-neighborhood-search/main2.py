import os
import time
from openpyxl import Workbook
from file_reader import read_txt_file
from file_writer import save_to_excel
from distance_finder import distance_matrix_generator, calculate_total_distance
from gap_calculator import read_lower_bounds, write_GAP_excel
from solution_interpreter import info_of_all_routes
from simulated_annealing import simulated_annealing_robust
from tabu import tabu_search_dynamic
from alns import alns_algorithm
from alns_operators import destroy_random, destroy_worst, destroy_route_removal, destroy_least_utilized
from alns_operators import repair_greedy, repair_regret, repair_savings
from vnd import vnd_algorithm



# Global parameters
alpha = 1
beta = 1000
initial_temperature = 300
cooling_rate = 0.95
tabu_tenure = 10
MAX_NO_IMPROVEMENT = 500  # Number of iterations without improvement


destroy_operators = [destroy_random, destroy_worst, destroy_route_removal, destroy_least_utilized]
repair_operators = [repair_greedy, repair_regret, repair_savings]


def get_initial_solution(method, nodes, Q, distances, initial_solution_path, sheet_name):
    initial_solution, _, _ = info_of_all_routes(initial_solution_path, sheet_name)
    for route_data in initial_solution:
        route_data['route_objects'] = [nodes[node_index] for node_index in route_data['route_indexes']]
    return [i['route_objects'] for i in initial_solution]

if __name__ == "__main__":
    initial_methods = ['constructive', 'GRASP', 'ACO']
    folder_name = '3-neighborhood-search'
    instances_directory_path = 'VRPTW Instances'
    LB_file_directory = 'VRPTW Instances/LB_VRPTW.xlsx'

    computation_times = {method: [] for method in initial_methods}
    all_18_routes = {method: [] for method in initial_methods}
    D = {method: [] for method in initial_methods}
    K = {method: [] for method in initial_methods}

    # Read lower bounds
    LB_K, LB_D = read_lower_bounds(LB_file_directory, 'Hoja1')

    for initial_method in initial_methods:
        print(f"\nProcessing for initial method: {initial_method}")

        # Create an Excel file for results
        results_excel_path = f"{folder_name}/results/VRPTW_tm_metaheuristic_ini_{initial_method}.xlsx"
        wb_results = Workbook()
        if "Sheet" in wb_results.sheetnames:
            wb_results.remove(wb_results["Sheet"])

        for sheet_number in range(1, 19):  # Iterate through all instances
            sheet_name = f'VRPTW{sheet_number}'
            instance_filename = f'{instances_directory_path}/{sheet_name}.txt'

            # Configure remaining time based on the instance
            if sheet_number >= 1 and sheet_number <= 6:
                remaining_time = 50e3
            elif sheet_number >= 7 and sheet_number <= 12:
                remaining_time = 200e3
            else:
                remaining_time = 750e3

            # Read input data
            n, Q, nodes = read_txt_file(instance_filename)
            distances = distance_matrix_generator(nodes)

            print(f" - Processing {sheet_name} with initial method: {initial_method}")
            start_time = time.time()
            initial_solution_path = f'{folder_name}/constructive-results/VRPTW_tm_{initial_method}.xlsx'

            # Get initial solution
            initial_solution = get_initial_solution(initial_method, nodes, Q, distances, initial_solution_path, sheet_name)

            # Apply Simulated Annealing
            routes = simulated_annealing_robust(
                initial_solution, distances, Q, initial_temperature, cooling_rate,
                remaining_time, start_time, alpha, beta
            )
            elapsed_time = time.time() - start_time
            remaining_time -= elapsed_time

            # Apply Tabu Search
            if remaining_time > 0:
                routes = tabu_search_dynamic(
                    routes, distances, Q, tabu_tenure, remaining_time, time.time(), alpha, beta
                )
                elapsed_time = time.time() - start_time
                remaining_time -= elapsed_time

            # Apply ALNS
            if remaining_time > 0:
                routes = alns_algorithm(
                    routes, distances, Q, destroy_operators, repair_operators,
                    remaining_time, time.time(), alpha, beta
                )
                elapsed_time = time.time() - start_time
                remaining_time -= elapsed_time

            # Apply VND
            if remaining_time > 0:
                routes = vnd_algorithm(routes, distances, Q, remaining_time, time.time())

            # Calculate total distance and number of routes
            total_distance = calculate_total_distance(routes, distances)
            route_count = len(routes)

            print(f"   - Total Distance: {total_distance}")
            print(f"   - Number of Routes: {route_count}")
            print(f"   - Remaining Time: {remaining_time:.2f} ms")

            # Save results
            computation_times[initial_method].append(elapsed_time)
            D[initial_method].append(total_distance)
            K[initial_method].append(route_count)
            all_18_routes[initial_method].append(routes)

            # Save sheet in results Excel file
            save_to_excel(wb_results, sheet_name, routes, total_distance, elapsed_time, distances)

        # Save Excel file after all instances
        wb_results.save(results_excel_path)

    # Create a single file for GAPs
    wb_gaps = Workbook()
    if "Sheet" in wb_gaps.sheetnames:
        wb_gaps.remove(wb_gaps["Sheet"])

    for method in initial_methods:
        sheet_name = f"{method}"
        write_GAP_excel(wb_gaps, LB_K, K[method], LB_D, D[method], computation_times[method], sheet_name)

    wb_gaps.save(f"{folder_name}/results/GAPs_for_metaheuristic_with_all_methods.xlsx")
